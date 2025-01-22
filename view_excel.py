from openpyxl import load_workbook

# Nhập đường dẫn file Excel
file_path = 'company_information.xlsx'

# Mở file Excel
try:
    workbook = load_workbook(file_path)
    sheet = workbook.active  # Chọn sheet đầu tiên

    # Lấy và hiển thị tất cả dữ liệu
    print("Dữ liệu trong file Excel:")
    for row in sheet.iter_rows(values_only=True):
        print(row)
except FileNotFoundError:
    print(f"File '{file_path}' không tồn tại.")
except Exception as e:
    print(f"Lỗi: {e}")


# xem danh sách các tài khoản để login vào trang hoadondientu
# python .\view_excel.py