import os
import argparse
from openpyxl import load_workbook

FILE_PATH = 'company_information.xlsx'
HOADON_COMPANY = "#"
HOADON_USERNAME = "#"

def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(description="Xóa dữ liệu trong file Excel.")
    parser.add_argument('--file-path', default=FILE_PATH, required=False,
                        help="Đường dẫn đến file Excel (mặc định: company_information.xlsx).")
    parser.add_argument('--company', default=HOADON_COMPANY, required=False, 
                        help="Tên công ty cần xóa.")
    parser.add_argument('--hoadon-username', default=HOADON_USERNAME, required=False, 
                        help="Username hóa đơn cần xóa.")
    return parser.parse_args()

def main():
    # Parse các tham số từ dòng lệnh hoặc biến môi trường
    args = parse_arguments()

    # Kiểm tra file tồn tại
    if not os.path.exists(args.file_path):
        print(f"File '{args.file_path}' không tồn tại. Vui lòng kiểm tra lại.")
        exit()

    # Mở file Excel
    workbook = load_workbook(args.file_path)
    sheet = workbook.active  # Chọn sheet đầu tiên

    # Duyệt qua các hàng và tìm hàng cần xóa
    row_to_delete = None
    for row in sheet.iter_rows(values_only=True):
        if (args.company == "#" or args.company == row[0]) and \
           (args.hoadon_username == "#" or args.hoadon_username == row[1]):
            row_to_delete = row
            break

    if row_to_delete:
        # Xóa hàng
        for i, row in enumerate(sheet.iter_rows(), start=1):
            if row[0].value == row_to_delete[0] and row[1].value == row_to_delete[1]:
                sheet.delete_rows(i)
                print(f"Đã xóa dòng: {row_to_delete}")
                break
    else:
        print("Không tìm thấy dữ liệu phù hợp để xóa.")

    # Lưu file
    workbook.save(args.file_path)
    print(f"Dữ liệu đã được cập nhật trong file '{args.file_path}'.")

if __name__ == "__main__":
    main()


# truyền tên công ty cần xóa [company]
# python .\delete_company_info_hoadon.py --company "cong ty mb"