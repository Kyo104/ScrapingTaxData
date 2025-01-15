import glob
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import pandas as pd
import os
import requests
from PIL import Image
from io import BytesIO
import base64
from sqlalchemy import create_engine, text
from sqlalchemy.exc import SQLAlchemyError


print('hello thuedientu')

# task 1 Đăng nhập vào website https://thuedientu.gdt.gov.vn/etaxnnt/Request
def initialize_driver():
    """Khởi tạo trình duyệt Chrome."""
    chrome_options = Options()
    chrome_options.add_argument("--headless=new") # for Chrome >= 109
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--force-device-scale-factor=1")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    
    driver = webdriver.Chrome(options=chrome_options)
    
    driver.maximize_window()  # Mở trình duyệt ở chế độ toàn màn hình
    time.sleep(2)
    return driver

# 1.1 Nhập username và password vào trang web 'thuedientu'
def login_to_thuedientu(driver, username, password):
    """Đăng nhập vào trang web 'thuedientu'."""
    url = 'https://thuedientu.gdt.gov.vn/etaxnnt/Request'
    driver.get(url)
    print('- Finish initializing a driver')
    time.sleep(2)

    # Nhấn nút Doanh Nghiệp
    doanh_nghiep_button = driver.find_element(By.XPATH, '//*[@id="bodyP"]/div[1]/div[4]/div/div[2]/div/div[2]/a')
    doanh_nghiep_button.click()
    time.sleep(3)
    print('- Finish Task 1: Login to Doanh_Nghiep')

    # Nhấn nút Đăng nhập
    login_button = driver.find_element(By.XPATH, '//*[@id="bodyP"]/div[1]/div[1]/div[3]/span[2]/button/strong/img')
    login_button.click()
    time.sleep(3)
    print('- Finish Task 1: Login to thuedientu')

    # Nhập tên đăng nhập
    username_field = driver.find_element(By.ID, '_userName')
    username_field.send_keys(username)
    print('- Finish keying in username_field')
    time.sleep(3)

    # Nhập mật khẩu
    password_field = driver.find_element(By.NAME, '_password')
    password_field.send_keys(password)
    print('- Finish keying in password_field')
    time.sleep(2)

    # Chọn đối tượng "Người nộp thuế"
    doi_tuong_select = driver.find_element(By.ID, 'login_type')
    select = Select(doi_tuong_select)
    select.select_by_value("01")
    print('- Finish keying in Doi_Tuong')
    time.sleep(2)
    
# Tải ảnh CAPTCHA về máy
def save_captcha_image(driver):
    """Tải ảnh CAPTCHA về máy."""
    try:
        # refresh_button = driver.find_element(By.CLASS_NAME, 'lam_moi_mxn')
        # refresh_button.click()
        # print("Refreshed CAPTCHA")

        # Sau đó, chụp lại CAPTCHA mới
        captcha_element = driver.find_element(By.ID, 'safecode')
        captcha_element.screenshot("captcha_image.png")
        print("[INFO] CAPTCHA đã được lưu tại captcha_image.png")
    except Exception as e:
        print(f"[ERROR] Lỗi khi lưu ảnh CAPTCHA: {e}")


API_KEY = "2fd46e197a886b8606fd140e3fe1f20b"  # Thay bằng API Key của bạn từ autocaptcha


# Gửi ảnh lên autocaptcha để giải mã
def solve_captcha(image_base64):
    """Gửi ảnh base64 lên autocaptcha và nhận mã CAPTCHA."""
    url = "https://autocaptcha.pro/api/captcha"
    payload = {
        "apikey": API_KEY,
        "img": image_base64,
        "type": 14  # Loại captcha, có thể cần thay đổi nếu không đúng
    }
    headers = {"Content-Type": "application/json"}

    try:
        # Gửi POST request
        response = requests.post(url, json=payload, headers=headers)

        # Kiểm tra nếu có lỗi trong phản hồi HTTP
        if response.status_code != 200:
            print(f"[ERROR] Error with request: {response.status_code}")
            print(f"[DEBUG] Response Text: {response.text}")
            return None

        # Phân tích phản hồi JSON
        response_data = response.json()

        # Kiểm tra xem API trả về thành công
        if response_data.get("success") and "captcha" in response_data:
            print(f"Mã captcha đã giải: {response_data['captcha']}")
            return response_data["captcha"]
        else:
            print(f"[ERROR] API response indicates failure: {response_data}")
            return None
    except Exception as e:
        print(f"[ERROR] Lỗi khi gửi yêu cầu giải CAPTCHA: {e}")
        return None


# Xử lý ảnh CAPTCHA và giải mã
def solve_captcha_from_file(file_path):
    """Đọc file CAPTCHA và gửi lên AntiCaptcha để giải mã."""
    try:
        # Đọc file captcha
        with open(file_path, 'rb') as file:
            img = Image.open(file)
            buffered = BytesIO()
            img.save(buffered, format="PNG")
            image_base64 = base64.b64encode(buffered.getvalue()).decode("utf-8")

        # Gửi ảnh base64 lên AntiCaptcha để giải mã
        captcha_text = solve_captcha(image_base64)

        # Chỉ trả về kết quả
        return captcha_text
    except Exception as e:
        print(f"[ERROR] Lỗi khi xử lý ảnh CAPTCHA: {e}")
        return None

    
      
    
# # 1.2 Nhập mã CAPTCHA tự động
# def enter_verification_code(driver, captcha_image_path):
#     """Giải mã CAPTCHA từ file và tự động nhập vào trường xác nhận."""
#     try:
#         # Giải mã CAPTCHA chỉ một lần
#         captcha_code = solve_captcha_from_file(captcha_image_path)
#         if not captcha_code:
#             print("[ERROR] Không thể giải mã CAPTCHA.")
#             return False

#         # Tìm trường nhập CAPTCHA
#         verification_code_field = driver.find_element(By.ID, 'vcode')

#         # Nhập mã CAPTCHA vào trường
#         verification_code_field.clear()
#         verification_code_field.send_keys(captcha_code)
#         time.sleep(2)

#         # Log giá trị sau khi nhập để kiểm tra
#         captcha_value = verification_code_field.get_attribute('value')
#         print(f"[INFO] CAPTCHA đã nhập: {captcha_value}")

#         return True
#     except Exception as e:
#         print(f"[ERROR] Lỗi khi nhập mã CAPTCHA: {e}")
#         return False
# ---------------------------------------------------------------------------------------------------------------------------------------------

# 1.2 Nhập mã captcha thủ công
def enter_verification_code(driver):
    """Nhập mã xác nhận."""
    # Yêu cầu người dùng nhập mã xác nhận
    code = input("Vui lòng nhập mã xác nhận: ")  # Người dùng tự nhập mã xác nhận
    # Tìm và nhập Mã xác nhận
    verification_code_field = driver.find_element(By.ID, 'vcode')
    verification_code_field.send_keys(code)
    print('- Finish keying in verification code')
    time.sleep(2)
    # Log giá trị sau khi nhập
    captcha_value = verification_code_field.get_attribute('value')
    print(f"[DEBUG] Giá trị Mã xác nhận sau khi nhập: {captcha_value}")

# ---------------------------------------------------------------------------------------------------------------------------------------------


def retry_user_pass_doituong(driver, username, password):
    # Nhập tên đăng nhập
    username_field = driver.find_element(By.ID, '_userName')
    username_field.send_keys(username)
    print('- Finish keying in username_field')
    time.sleep(3)

    # Nhập mật khẩu
    password_field = driver.find_element(By.NAME, '_password')
    password_field.send_keys(password)
    print('- Finish keying in password_field')
    time.sleep(2)

    # Chọn đối tượng "Người nộp thuế"
    doi_tuong_select = driver.find_element(By.ID, 'login_type')
    select = Select(doi_tuong_select)
    select.select_by_value("01")
    print('- Finish keying in Doi_Tuong')
    time.sleep(2)
    
    
# 1.3 Nhấn nút đăng nhập sau cùng hoàn tất việc login vào trang web
def submit_form(driver, username, password, captcha_image_path):
    """Nhấn nút để hoàn tất đăng nhập."""
    try:
        while True:
            # Nhấn nút để gửi biểu mẫu
            submit_button = driver.find_element(By.XPATH, '//*[@id="dangnhap"]')
            submit_button.click()
            print('- Finish submitting the form')
            
            # Kiểm tra nếu có thông báo lỗi CAPTCHA
            try:
                # Chờ thông báo lỗi CAPTCHA
                error_message = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "Mã xác thực không chính xác")]'))
                )
                if error_message:
                    print("[ERROR] Mã xác nhận nhập sai. Đang thử lại...")
                    # Nhập lại các trường thông tin
                    retry_user_pass_doituong(driver, username, password)
                    #---------------------------------------------------------------------------------------------------------------------------------
                    # Lưu và giải mã CAPTCHA mới
                    save_captcha_image(driver)
                    
                    enter_verification_code(driver) # thủ công
                    # enter_verification_code(driver, captcha_image_path) # tự đông nhập mã captcha
                    continue  # Thử lại
            except TimeoutException:
                print("[DEBUG] Mã xác nhận được xác thực thành công")
            
            # Kiểm tra nếu đăng nhập thành công
            try:
                # Chờ thẻ div có id "ddtabs1" xuất hiện
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.ID, "ddtabs1"))
                )
                # Tìm trong ul có id "tabmenu" và kiểm tra thẻ span với text "Tra cứu"
                tra_cuu_element = driver.find_element(
                    By.XPATH, '//div[@id="ddtabs1"]//ul[@id="tabmenu"]//li//a//span[text()="Tra cứu"]'
                )
                if tra_cuu_element:
                    print("[INFO] Đăng nhập thành công! Đã vào trang chính.")
                    return  # Thoát khỏi hàm khi thành công
            except TimeoutException:
                print("[DEBUG] Không tìm thấy dấu hiệu đăng nhập thành công. Thử lại...")
                continue  # Thử lại nếu không tìm thấy dấu hiệu thành công
            
            # Nếu không vào được vòng lặp, thoát ra
            break
    except Exception as e:
        print(f"Đã xảy ra lỗi khi nhấn nút submit: {e}")

    
    
    
# Task 2 crawl dữ liệu ở tab Truy vấn và xuất file xlsx lưu vào máy
# ( Hàm Thêm stt sau mỗi file trùng tên )
def get_unique_filename(base_filename):
    """
    Tạo tên file duy nhất nếu file đã tồn tại, bằng cách thêm số thứ tự theo định dạng (1), (2),...
    """
    if not os.path.exists(base_filename):
        return base_filename

    base, ext = os.path.splitext(base_filename)
    counter = 1
    new_filename = f"{base} ({counter}){ext}"

    while os.path.exists(new_filename):
        counter += 1
        new_filename = f"{base} ({counter}){ext}"

    return new_filename

# ( Hàm lưu dữ liệu vào file Excel theo form đã chỉnh )
def save_to_excel_with_style(df, file_name):
    """Lưu dữ liệu vào file Excel với tiêu đề màu xanh và khung viền."""
    # Tạo tên file duy nhất nếu cần
    unique_file_name = get_unique_filename(file_name)
    # Tạo workbook và sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    # Thêm tiêu đề
    title_fill = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid")  # Màu xanh
    title_font = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                    top=Side(style="thin"), bottom=Side(style="thin"))
    
    # Thêm dữ liệu
    for col_idx, column_name in enumerate(df.columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=column_name)
        cell.fill = title_fill
        cell.font = title_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm các hàng dữ liệu
    for row_idx, row_data in enumerate(df.values, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Lưu file
    workbook.save(unique_file_name)
    print(f"Dữ liệu đã được lưu vào file Excel: {file_name}")
    # Trả về tên file để điều chỉnh kích thước cột
    return unique_file_name
    
# ( Hàm lưu dữ liệu vào file Excel theo form đã chỉnh độ rộng của từng cột )
def adjust_column_width(file_path):
    # Mở file Excel đã lưu
    workbook = load_workbook(file_path)
    sheet = workbook.active  # Lấy sheet đầu tiên (hoặc tên cụ thể nếu cần)

    # Duyệt qua các cột để tự động điều chỉnh độ rộng
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Lấy tên cột (A, B, C,...)

        # Tính độ dài lớn nhất của nội dung trong cột
        for cell in column:
            try:
                if cell.value:  # Bỏ qua ô trống
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        # Đặt độ rộng cột dựa trên độ dài lớn nhất
        adjusted_width = max_length + 2  # Thêm khoảng trống
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Lưu file sau khi chỉnh sửa
    workbook.save(file_path)
    print(f"Đã tự động điều chỉnh độ rộng cột trong file {file_path}")

# 2.1 Chọn vào mục tra cứu thuedientu
def crawl(driver):
    # Nhấn nút tra cứu
    tra_cuu_button = driver.find_element(By.XPATH, '//*[@id="tabmenu"]/li[5]/a')
    tra_cuu_button.click()
    print('- Finish click tra cuu')
    time.sleep(3)

    # Kiểm tra nếu nút "Truy vấn" nằm trong iframe
    try:
        iframe = driver.find_element(By.XPATH, '//*[@id="tranFrame"]')  # Thay 'iframe_id' nếu cần
        driver.switch_to.frame(iframe)
        print("- Đã chuyển vào iframe")
    except NoSuchElementException:
        print("- Không tìm thấy iframe, tiếp tục thao tác trên trang chính.")

    # Đợi phần tử hiển thị và click bằng JavaScript
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[value="Truy vấn"]')))
    driver.execute_script("document.querySelector('.button_vuong.awesome').click();")

    print('- Finish click Truy van')

    # Quay lại trang chính nếu đã vào iframe
    driver.switch_to.default_content()
    time.sleep(5)
    
    # Bước 1: Lấy mã nguồn HTML của trang hiện tại
    website_url = driver.current_url
    print(f"URL hiện tại sau khi truy vấn: {website_url}")
    
    # Sử dụng `driver.page_source` mà không cần `.text`
    page_source = driver.page_source
    
    # In ra một phần mã nguồn để kiểm tra
    print("Mã nguồn HTML của trang sau khi truy vấn:")
    print(page_source[:1000])  # In ra 1000 ký tự đầu tiên của mã nguồn

    # Phân tích HTML bằng BeautifulSoup
    soup = BeautifulSoup(page_source, 'lxml')

    # Bước 2: Tìm bảng có id là 'data_content_onday'
    table = soup.find('table', id='data_content_onday')

    # Kiểm tra nếu không tìm thấy bảng
    if table is None:
        print("Không tìm thấy bảng với id 'data_content_onday'.")
        # Kiểm tra nếu bảng có thể nằm trong một iframe khác
        iframe_elements = driver.find_elements(By.TAG_NAME, 'iframe')
        print(f"Found {len(iframe_elements)} iframe(s) on the page.")
        for i, iframe in enumerate(iframe_elements):
            print(f"Switching to iframe {i+1}")
            driver.switch_to.frame(iframe)
            time.sleep(3)  # Đảm bảo iframe đã tải xong
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'lxml')
            table = soup.find('table', id='data_content_onday')
            if table:
                print(f"Found the table in iframe {i+1}")
                break  # Nếu tìm thấy bảng, thoát khỏi vòng lặp
            driver.switch_to.default_content()  # Quay lại trang chính nếu không tìm thấy trong iframe
    else:
        print("Bảng đã được tìm thấy.")

    # Nếu bảng được tìm thấy, tiếp tục xử lý dữ liệu
    if table:
        elements = []
        rows = table.find_all('tr')  # Lấy tất cả các dòng trong bảng

        for row in rows[2:]:  # Bỏ qua 2 dòng đầu tiên
            cells = row.find_all('td')  # Lấy tất cả các ô dữ liệu trong một dòng
            row_data = [cell.get_text(strip=True) for cell in cells]  # Lấy văn bản trong ô
            elements.append(row_data)

        # Lấy tiêu đề cột từ bảng
        world_titles = table.find_all('span')
        world_table_titles = [title.text.strip() for title in world_titles]

        # Chỉ giữ các tiêu đề và dữ liệu đến "Tính chất khoản nộp"
        try:
            cutoff_index = world_table_titles.index("Tính chất khoản nộp") + 1
            world_table_titles = world_table_titles[:cutoff_index]
            elements = [row[:cutoff_index] for row in elements]
        except ValueError:
            print("Không tìm thấy cột 'Tính chất khoản nộp' trong tiêu đề.")
            cutoff_index = len(world_table_titles)  # Giữ toàn bộ nếu không tìm thấy

        # Chuyển dữ liệu thành DataFrame
        df = pd.DataFrame(elements, columns=world_table_titles)

    else:
        print("Không tìm thấy bảng với id 'data_content_onday'.")
        df = pd.DataFrame()  # Trả về DataFrame rỗng nếu không tìm thấy bảng

    return df
           
        

# 2.2 Lưu và xứ lý dữ liệu vào database PostgreSQL
# Hàm tạo và kết nối đến database PostgreSQL
def create_and_connect_to_database(db_name, user, password, host='localhost', port='5432'):
    """Tạo một database mới nếu chưa tồn tại và kết nối đến nó."""
    try:
        # Kết nối đến PostgreSQL (mặc định là database hệ thống 'postgres')
        engine = create_engine(f'postgresql://{user}:{password}@{host}:{port}', isolation_level='AUTOCOMMIT')

        # Tạo database nếu chưa tồn tại
        with engine.connect() as conn:
            result = conn.execute(
                text("SELECT 1 FROM pg_catalog.pg_database WHERE datname = :db_name"),
                {"db_name": db_name}
            )
            exists = result.fetchone()
            if not exists:
                conn.execute(text(f"CREATE DATABASE {db_name}"))
                print(f"Database '{db_name}' đã được tạo.")
            else:
                print(f"Database '{db_name}' đã tồn tại.")

        # Kết nối đến database vừa tạo
        engine = create_engine(f'postgresql://{user}:{password}@{host}:{port}/{db_name}')
        print(f"Kết nối thành công đến database: {db_name}")
        return engine

    except Exception as e:
        print(f"Lỗi khi tạo hoặc kết nối đến database: {e}")
        return None


# Hàm đọc file Excel và tải dữ liệu lên PostgreSQL
def upload_excel_to_postgres(db_config):
    try:
        # Lấy file Excel mới nhất từ thư mục
        list_of_files = glob.glob(f"./data_thue_dien_tu.xlsx")
        if not list_of_files:
            print("Không tìm thấy file Excel nào trong thư mục.")
            return

        latest_file = max(list_of_files, key=os.path.getctime)  # Lấy file mới nhất
        print(f"File Excel mới nhất: {latest_file}")

        # Kết nối tới PostgreSQL
        engine = create_engine(
            f"postgresql+psycopg2://{db_config['user']}:{db_config['password']}@{db_config['host']}:{db_config['port']}/{db_config['database']}"
        )
        connection = engine.connect()

        # Đọc file Excel
        data = pd.read_excel(latest_file)

        # Xử lý dữ liệu rỗng (thay None hoặc NaN bằng "")
        data = data.fillna("")

        # Tải dữ liệu lên PostgreSQL vào bảng tạm (tạo nếu chưa tồn tại)
        data.to_sql('tax_data', con=engine, if_exists='replace', index=False)
        print("Đã tải dữ liệu lên bảng tax_data.")

        # Thêm cột tax_data_id làm khóa chính (serial)
        with engine.begin() as conn:
            conn.execute(text("""ALTER TABLE tax_data ADD COLUMN tax_data_id SERIAL PRIMARY KEY;"""))
        print("Đã thêm cột tax_data_id làm khóa chính cho bảng tax_data.")

    except Exception as e:
        print(f"Lỗi khi tải dữ liệu lên PostgreSQL: {e}")


def process_and_create_tables(db_config):
    try:
        engine = create_engine(
            f"postgresql+psycopg2://{db_config['user']}:{db_config['password']}@{db_config['host']}:{db_config['port']}/{db_config['database']}"
        )
        tax_data = pd.read_sql("SELECT * FROM tax_data", con=engine)

        # Duyệt lần 1: Lọc và tạo bảng phụ với mục A, B, C
        sections = ['A', 'B', 'C']
        for i, section in enumerate(sections):
            start_index = tax_data[tax_data['STT'].str.startswith(f"{section}.")].index[0]
            if i < len(sections) - 1:
                next_section = sections[i + 1]
                end_index = tax_data[tax_data['STT'].str.startswith(f"{next_section}.")].index[0]
                section_data = tax_data.iloc[start_index:end_index]
            else:
                section_data = tax_data.iloc[start_index:]
            section_table_name = f"tax_data_{section.lower()}"
            section_data.to_sql(section_table_name, con=engine, if_exists='replace', index=False)
            with engine.begin() as conn:
                conn.execute(text(f"""
                    ALTER TABLE {section_table_name} ADD COLUMN {section_table_name}_id SERIAL PRIMARY KEY;
                    ALTER TABLE {section_table_name} ADD CONSTRAINT fk_tax_data FOREIGN KEY (tax_data_id) REFERENCES tax_data (tax_data_id);
                """))
            print(f"Đã tạo bảng phụ {section_table_name} với {len(section_data)} hàng.")

        # Duyệt lần 2: Lọc và tạo bảng con từ từng bảng phụ
        for section in sections:
            section_table_name = f"tax_data_{section.lower()}"
            section_data = pd.read_sql(f"SELECT * FROM {section_table_name}", con=engine)
            subsections = section_data['STT'][section_data['STT'].str.match(r'^(I|II|III|IV|V|VI|VII|VIII|IX|X)\..*')].tolist()
            if not subsections:
                print(f"Bảng phụ {section_table_name} không có mục con, bỏ qua.")
                continue
            for j, subsection in enumerate(subsections):
                start_index = section_data[section_data['STT'] == subsection].index[0]
                if j < len(subsections) - 1:
                    end_index = section_data[section_data['STT'] == subsections[j + 1]].index[0]
                    subsection_data = section_data.iloc[start_index:end_index]
                else:
                    subsection_data = section_data.iloc[start_index:]
                subsection_table_name = f"{section_table_name}_{subsection.split('.')[0].lower()}"
                subsection_data.to_sql(subsection_table_name, con=engine, if_exists='replace', index=False)
                with engine.begin() as conn:
                    conn.execute(text(f"""
                        ALTER TABLE {subsection_table_name} ADD COLUMN {subsection_table_name}_id SERIAL PRIMARY KEY;
                        ALTER TABLE {subsection_table_name} ADD CONSTRAINT fk_{section_table_name} FOREIGN KEY ({section_table_name}_id) REFERENCES {section_table_name} ({section_table_name}_id);
                        ALTER TABLE {subsection_table_name} ADD CONSTRAINT fk_tax_data FOREIGN KEY (tax_data_id) REFERENCES tax_data (tax_data_id);
                    """))
                print(f"Đã tạo bảng con {subsection_table_name} với {len(subsection_data)} hàng.")
    except Exception as e:
        print(f"Lỗi khi xử lý và tạo bảng: {e}")



    
def main():
    """Chạy tất cả các bước trong quy trình đăng nhập."""
    driver = initialize_driver()

    # Thay thế username, password vào
    username = "0101652097-ql"
    password = "At2025$$$"
    
    captcha_image_path = "captcha_image.png"

    # Định nghĩa cấu hình kết nối database
    db_config = {
        'host': 'localhost',
        'port': 5432,
        'user': 'postgres',
        'password': '123456',
        'database': 'taxdata'
    }

    # Tạo và kết nối đến database
    engine = create_and_connect_to_database(
        db_name=db_config['database'],
        user=db_config['user'],
        password=db_config['password'],
        host=db_config['host'],
        port=db_config['port']
    )

    
    try:
        login_to_thuedientu(driver, username, password)
        
        save_captcha_image(driver)
        
        enter_verification_code(driver) # thủ công
        # enter_verification_code(driver, captcha_image_path) # tự động
        
        
        submit_form(driver, username, password, captcha_image_path)
        df = crawl(driver)
        
        
        if engine:
            # Lưu dữ liệu vào file Excel
            file_path = 'data_thue_dien_tu.xlsx'  # Thay bằng tên file của bạn
            unique_file_name = save_to_excel_with_style(df, file_path)   
            # Gọi hàm điều chỉnh kích thước cột
            adjust_column_width(unique_file_name)  

            # Gọi hàm tải dữ liệu
            upload_excel_to_postgres(db_config)

            # Gọi hàm xử lý và tạo bảng
            process_and_create_tables(db_config)
            

        
    except Exception as e:
        print(f"An error occurred: {e}")
#     finally:
#         driver.quit()  # Đóng trình duyệt sau khi hoàn thành

if __name__ == '__main__':
    main()