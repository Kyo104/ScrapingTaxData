import glob
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import pandas as pd
import os
import requests
from PIL import Image
from io import BytesIO
import base64
from sqlalchemy import create_engine, text
from sqlalchemy.exc import SQLAlchemyError
import argparse
import requests
import json

# =================== BIẾN MÔI TRƯỜNG ===================

# API key cho dịch vụ giải captcha
API_KEY = "4b9744cc99fd188fb23d1440fbc45639"  

# Mục thông tin kết nối database
DB_USER = "postgres" 
DB_PASSWORD = "123456"  
DB_NAME = "crawling_data"
DB_HOST = "localhost"  
DB_PORT = "5432"  


# URL Webhook Slack mặc định
WEBHOOK_URL = 'https://hooks.slack.com/services/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'

print('hello thuedientu')
# ==============================================================================

def parse_arguments():
    parser = argparse.ArgumentParser(description='Thuế Điện Tử Data Crawler')
    
    parser.add_argument('--api-key', default=API_KEY,
                       help='API key từ trang web autocaptcha để giải captcha')
    parser.add_argument('--db-user', default=DB_USER,
                       help='PostgreSQL username')
    parser.add_argument('--db-password', default=DB_PASSWORD,
                       help='PostgreSQL password')
    parser.add_argument('--db-name', default=DB_NAME,
                       help='Database name')
    parser.add_argument('--db-host', default=DB_HOST,
                       help='Database host')
    parser.add_argument('--db-port', default=DB_PORT,
                       help='Database port')
    parser.add_argument('--webhook-url', default=WEBHOOK_URL, required=False,
                        help='Liên kết webhook từ Slack')  # Thêm dòng này
    
    return parser.parse_args()

args = parse_arguments()
webhook_url = args.webhook_url
print(f"Sử dụng webhook_url: {webhook_url}")

# task 1 Đăng nhập vào website https://thuedientu.gdt.gov.vn/etaxnnt/Request
def initialize_driver():
    """Khởi tạo trình duyệt Chrome."""
    chrome_options = Options()
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--headless=new") # for Chrome >= 109
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--force-device-scale-factor=1")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    
    driver = webdriver.Chrome(options=chrome_options)
    
    driver.maximize_window()  # Mở trình duyệt ở chế độ toàn màn hình
    time.sleep(2)
    send_slack_notification('======== Workflow ThueDienTu ==========.', webhook_url)
    return driver

# 1.1 Nhập username và password vào trang web 'thuedientu'
def login_to_thuedientu(driver, username, password, company):
    """Đăng nhập vào trang web 'thuedientu'."""
    url = 'https://thuedientu.gdt.gov.vn/etaxnnt/Request'
    driver.get(url)
    print('- Finish initializing a driver')
    time.sleep(2)
    
    print(f'- Đang đăng nhập cho công ty: {company}')
    send_slack_notification(f'[INFO] Chương trình đang login vào công ty: {company}', webhook_url)
    
    # Nhấn nút Doanh Nghiệp
    doanh_nghiep_button = driver.find_element(By.XPATH, '//*[@id="bodyP"]/div[1]/div[4]/div/div[2]/div/div[2]/a')
    doanh_nghiep_button.click()
    time.sleep(3)
    print('- Finish Task 1: Login to Doanh_Nghiep')

    # Nhấn nút Đăng nhập
    login_button = driver.find_element(By.XPATH, '//*[@id="bodyP"]/div[1]/div[1]/div[3]/span[2]/button/strong/img')
    login_button.click()
    time.sleep(3)
    print('- Finish Task 1: Login to thuedientu')
    
    # click vào Thue dien tu
    btn_tk_thue = driver.find_element(By.XPATH, '//*[@id="icon-1"]')
    btn_tk_thue.click()
    time.sleep(3)
    print('- Finish Task 1: Login to icon')

    # Nhập tên đăng nhập
    username_field = driver.find_element(By.ID, '_userName')
    username_field.send_keys(username)
    print('- Finish keying in username_field')
    time.sleep(3)

    # Nhập mật khẩu
    password_field = driver.find_element(By.NAME, '_password')
    password_field.send_keys(password)
    print('- Finish keying in password_field')
    time.sleep(2)

    # Chọn đối tượng "Người nộp thuế"
    doi_tuong_select = driver.find_element(By.ID, 'login_type')
    select = Select(doi_tuong_select)
    select.select_by_value("01")
    print('- Finish keying in Doi_Tuong')
    time.sleep(2)

def send_slack_notification(message, webhook_url):
    headers = {
        'Content-Type': 'application/json',
    }
    payload = {
        "text": message  # Nội dung thông báo
    }
    try:  
      response = requests.post(webhook_url, headers=headers, data=json.dumps(payload))
      
      if response.status_code == 200:
            print("Thông báo đã được gửi thành công!")
      else:
            print(f"Lỗi khi gửi thông báo: {response.status_code}, {response.text}")
    except:
          pass

# Tải ảnh CAPTCHA về máy
def save_captcha_image(driver):
    """Tải ảnh CAPTCHA về máy."""
    try:
        # Sau đó, chụp lại CAPTCHA mới
        captcha_element = driver.find_element(By.ID, 'safecode')
        captcha_element.screenshot("captcha_image.png")
        print("[INFO] CAPTCHA đã được lưu tại captcha_image.png")
    except Exception as e:
        print(f"[ERROR] Lỗi khi lưu ảnh CAPTCHA: {e}")
        send_slack_notification('[ERROR] Chương trình chạy thất bại', webhook_url)

# Gửi ảnh lên autocaptcha để giải mã
def solve_captcha(image_base64):
    """Gửi ảnh base64 lên autocaptcha và nhận mã CAPTCHA."""
    url = "https://autocaptcha.pro/api/captcha"
    payload = {
        "apikey": API_KEY,
        "img": image_base64,
        "type": 14  # Loại captcha, có thể cần thay đổi nếu không đúng
    }
    headers = {"Content-Type": "application/json"}

    try:
        # Gửi POST request
        response = requests.post(url, json=payload, headers=headers)

        # Kiểm tra nếu có lỗi trong phản hồi HTTP
        if response.status_code != 200:
            print(f"[ERROR] Error with request: {response.status_code}")
            print(f"[DEBUG] Response Text: {response.text}")
            return None

        # Phân tích phản hồi JSON
        response_data = response.json()

        # Kiểm tra xem API trả về thành công
        if response_data.get("success") and "captcha" in response_data:
            print(f"Mã captcha đã giải: {response_data['captcha']}")
            return response_data["captcha"]
        else:
            print(f"[ERROR] API response indicates failure: {response_data}")
            send_slack_notification(f'[ERROR] Chương trình chạy thất bại {response_data}', webhook_url)
            return None
    except Exception as e:
        print(f"[ERROR] Lỗi khi gửi yêu cầu giải CAPTCHA: {e}")
        send_slack_notification('[ERROR] Chương trình chạy thất bại', webhook_url)
        return None

# Xử lý ảnh CAPTCHA và giải mã
def solve_captcha_from_file(file_path):
    """Đọc file CAPTCHA và gửi lên AntiCaptcha để giải mã."""
    try:
        # Đọc file captcha
        with open(file_path, 'rb') as file:
            img = Image.open(file)
            buffered = BytesIO()
            img.save(buffered, format="PNG")
            image_base64 = base64.b64encode(buffered.getvalue()).decode("utf-8")

        # Gửi ảnh base64 lên AntiCaptcha để giải mã
        captcha_text = solve_captcha(image_base64)

        # Chỉ trả về kết quả
        return captcha_text
    except Exception as e:
        print(f"[ERROR] Lỗi khi xử lý ảnh CAPTCHA: {e}")
        send_slack_notification('[ERROR] Chương trình chạy thất bại', webhook_url)
        return None

# 1.2 Nhập mã CAPTCHA tự động
def enter_verification_code(driver, captcha_image_path):
    """Giải mã CAPTCHA từ file và tự động nhập vào trường xác nhận."""
    try:
        # Giải mã CAPTCHA chỉ một lần
        captcha_code = solve_captcha_from_file(captcha_image_path)
        if not captcha_code:
            print("[ERROR] Không thể giải mã CAPTCHA.")
            return False

        # Tìm trường nhập CAPTCHA
        verification_code_field = driver.find_element(By.ID, 'vcode')

        # Nhập mã CAPTCHA vào trường
        verification_code_field.clear()
        verification_code_field.send_keys(captcha_code)
        time.sleep(2)

        # Log giá trị sau khi nhập để kiểm tra
        captcha_value = verification_code_field.get_attribute('value')
        print(f"[INFO] CAPTCHA đã nhập: {captcha_value}")

        return True
    except Exception as e:
        print(f"[ERROR] Lỗi khi nhập mã CAPTCHA: {e}")
        send_slack_notification('[ERROR] Chương trình chạy thất bại', webhook_url)
        return False


def retry_user_pass_doituong(driver, username, password):
    # Nhập tên đăng nhập
    username_field = driver.find_element(By.ID, '_userName')
    username_field.send_keys(username)
    print('- Finish keying in username_field')
    time.sleep(3)

    # Nhập mật khẩu
    password_field = driver.find_element(By.NAME, '_password')
    password_field.send_keys(password)
    print('- Finish keying in password_field')
    time.sleep(2)

    # Chọn đối tượng "Người nộp thuế"
    doi_tuong_select = driver.find_element(By.ID, 'login_type')
    select = Select(doi_tuong_select)
    select.select_by_value("01")
    print('- Finish keying in Doi_Tuong')
    time.sleep(2)
    
# 1.3 Nhấn nút đăng nhập sau cùng hoàn tất việc login vào trang web
def submit_form(driver, username, password, captcha_image_path):
    """Nhấn nút để hoàn tất đăng nhập."""
    try:
        attempt = 0  # Biến theo dõi số lần thử đăng nhập
        while True:
            
            attempt += 1  # Tăng số lần thử đăng nhập
            # Nhấn nút để gửi biểu mẫu
            submit_button = driver.find_element(By.XPATH, '//*[@id="dangnhap"]')
            submit_button.click()
            print(f'- Finish submitting the form (attempt {attempt})')
            send_slack_notification(f'[INFO] Chương trình đang thực hiên login lần {attempt}', webhook_url)
            
            # Kiểm tra nếu có thông báo lỗi CAPTCHA
            try:
                # Chờ thông báo lỗi CAPTCHA
                error_message = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "Mã xác thực không chính xác")]'))
                )
                if error_message:
                    print("[ERROR] Mã xác nhận nhập sai. Đang thử lại...")
                    send_slack_notification('[ERROR] Login thất bại, đang thử lại', webhook_url)
                    # Nhập lại các trường thông tin
                    retry_user_pass_doituong(driver, username, password)
                    
                    # Lưu và giải mã CAPTCHA mới
                    save_captcha_image(driver)
                    
                    # enter_verification_code(driver) # thủ công
                    enter_verification_code(driver, captcha_image_path) # tự đông nhập mã captcha
                    continue  # Thử lại
                
            except TimeoutException:
                print("[DEBUG] Mã xác nhận được xác thực thành công")
            
            # Kiểm tra nếu đăng nhập thành công
            try:
                # Chờ thẻ div có id "ddtabs1" xuất hiện
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.ID, "ddtabs1"))
                )
                # Tìm trong ul có id "tabmenu" và kiểm tra thẻ span với text 
                tra_cuu_element = driver.find_element(
                    By.XPATH, '//div[@id="ddtabs1"]//ul[@id="tabmenu"]//li//a//span[text()="Tra cứu"]'
                )
                if tra_cuu_element:
                    print("[INFO] Đăng nhập thành công! Đã vào trang chính.")
                    send_slack_notification('[SUCCESS] Đăng nhập thành công! Đã vào trang chính.', webhook_url)
                    return  # Thoát khỏi hàm khi thành công
            except TimeoutException:
                print("[DEBUG] Không tìm thấy dấu hiệu đăng nhập thành công. Thử lại...")
                continue  # Thử lại nếu không tìm thấy dấu hiệu thành công
            
            # Nếu không vào được vòng lặp, thoát ra
            break
    except Exception as e:
        print(f"Đã xảy ra lỗi khi nhấn nút submit: {e}")
        send_slack_notification('[ERROR] Chương trình chạy thất bại', webhook_url)
    
# Task 2 crawl dữ liệu ở tab Truy vấn và xuất file xlsx lưu vào máy
def get_unique_filename(base_filename):
    """
    Tạo tên file duy nhất nếu file đã tồn tại, bằng cách thêm số thứ tự theo định dạng (1), (2),...
    """
    if not os.path.exists(base_filename):
        return base_filename

    base, ext = os.path.splitext(base_filename)
    counter = 1
    new_filename = f"{base} ({counter}){ext}"

    while os.path.exists(new_filename):
        counter += 1
        new_filename = f"{base} ({counter}){ext}"

    return new_filename

# ( Hàm lưu dữ liệu vào file Excel theo form đã chỉnh )
def save_to_excel_with_style(df, file_name):
    """Lưu dữ liệu vào file Excel với tiêu đề màu xanh và khung viền."""
    # Tạo tên file duy nhất nếu cần
    unique_file_name = get_unique_filename(file_name)
    # Tạo workbook và sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    # Thêm tiêu đề
    title_fill = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid")  # Màu xanh
    title_font = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                    top=Side(style="thin"), bottom=Side(style="thin"))
    
    # Thêm dữ liệu
    for col_idx, column_name in enumerate(df.columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=column_name)
        cell.fill = title_fill
        cell.font = title_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm các hàng dữ liệu
    for row_idx, row_data in enumerate(df.values, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Lưu file
    workbook.save(unique_file_name)
    print(f"Dữ liệu đã được lưu vào file Excel: {file_name}")
    # Trả về tên file để điều chỉnh kích thước cột
    return unique_file_name
    
# ( Hàm lưu dữ liệu vào file Excel theo form đã chỉnh độ rộng của từng cột )
def adjust_column_width(file_path):
    # Mở file Excel đã lưu
    workbook = load_workbook(file_path)
    sheet = workbook.active  # Lấy sheet đầu tiên (hoặc tên cụ thể nếu cần)

    # Duyệt qua các cột để tự động điều chỉnh độ rộng
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Lấy tên cột (A, B, C,...)

        # Tính độ dài lớn nhất của nội dung trong cột
        for cell in column:
            try:
                if cell.value:  # Bỏ qua ô trống
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        # Đặt độ rộng cột dựa trên độ dài lớn nhất
        adjusted_width = max_length + 2  # Thêm khoảng trống
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Lưu file sau khi chỉnh sửa
    workbook.save(file_path)
    print(f"Đã tự động điều chỉnh độ rộng cột trong file {file_path}")

# 2.1 Chọn vào mục tra cứu thuedientu
def crawl(driver):
    # Nhấn nút tra cứu
    tra_cuu_button = driver.find_element(By.XPATH, '//*[@id="tabmenu"]/li[5]/a')
    tra_cuu_button.click()
    print('- Finish click tra cuu')
    time.sleep(3)

    # Kiểm tra nếu nút "Truy vấn" nằm trong iframe
    try:
        iframe = driver.find_element(By.XPATH, '//*[@id="tranFrame"]')  # Thay 'iframe_id' nếu cần
        driver.switch_to.frame(iframe)
        print("- Đã chuyển vào iframe")
    except NoSuchElementException:
        print("- Không tìm thấy iframe, tiếp tục thao tác trên trang chính.")

    # Đợi phần tử hiển thị và click bằng JavaScript
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[value="Truy vấn"]')))
    driver.execute_script("document.querySelector('.button_vuong.awesome').click();")

    print('- Finish click Truy van')

    # Quay lại trang chính nếu đã vào iframe
    driver.switch_to.default_content()
    time.sleep(5)
    
    # Bước 1: Lấy mã nguồn HTML của trang hiện tại
    website_url = driver.current_url
    print(f"URL hiện tại sau khi truy vấn: {website_url}")
    
    # Sử dụng `driver.page_source` mà không cần `.text`
    page_source = driver.page_source
    
    # In ra một phần mã nguồn để kiểm tra
    print("Mã nguồn HTML của trang sau khi truy vấn:")
    print(page_source[:1000])  # In ra 1000 ký tự đầu tiên của mã nguồn

    # Phân tích HTML bằng BeautifulSoup
    soup = BeautifulSoup(page_source, 'lxml')

    # Bước 2: Tìm bảng có id là 'data_content_onday'
    table = soup.find('table', id='data_content_onday')

    # Kiểm tra nếu không tìm thấy bảng
    if table is None:
        print("Không tìm thấy bảng với id 'data_content_onday'.")
        # Kiểm tra nếu bảng có thể nằm trong một iframe khác
        iframe_elements = driver.find_elements(By.TAG_NAME, 'iframe')
        print(f"Found {len(iframe_elements)} iframe(s) on the page.")
        for i, iframe in enumerate(iframe_elements):
            print(f"Switching to iframe {i+1}")
            driver.switch_to.frame(iframe)
            time.sleep(3)  # Đảm bảo iframe đã tải xong
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'lxml')
            table = soup.find('table', id='data_content_onday')
            if table:
                print(f"Found the table in iframe {i+1}")
                break  # Nếu tìm thấy bảng, thoát khỏi vòng lặp
            driver.switch_to.default_content()  # Quay lại trang chính nếu không tìm thấy trong iframe
    else:
        print("Bảng đã được tìm thấy.")

    # Nếu bảng được tìm thấy, tiếp tục xử lý dữ liệu
    if table:
        elements = []
        rows = table.find_all('tr')  # Lấy tất cả các dòng trong bảng

        for row in rows[2:]:  # Bỏ qua 2 dòng đầu tiên
            cells = row.find_all('td')  # Lấy tất cả các ô dữ liệu trong một dòng
            row_data = [cell.get_text(strip=True) for cell in cells]

            if len(row_data) > 1:  # Đảm bảo có dữ liệu và có ít nhất 2 cột (tránh lỗi index)
                elements.append(row_data[1:])  # Bỏ cột đầu tiên (STT)

        # Lấy tiêu đề cột từ bảng
        world_titles = table.find_all('span')
        world_table_titles = [title.text.strip() for title in world_titles]

        if len(world_table_titles) > 1:
            world_table_titles = world_table_titles[1:]  # Bỏ cột STT trong tiêu đề

        # Chỉ giữ các tiêu đề và dữ liệu đến "Tính chất khoản nộp"
        try:
            cutoff_index = world_table_titles.index("Tính chất khoản nộp") + 1
            world_table_titles = world_table_titles[:cutoff_index]
            elements = [row[:cutoff_index] for row in elements]
        except ValueError:
            print("Không tìm thấy cột 'Tính chất khoản nộp' trong tiêu đề.")
            cutoff_index = len(world_table_titles)  # Giữ toàn bộ nếu không tìm thấy

        # Chuyển dữ liệu thành DataFrame
        df = pd.DataFrame(elements, columns=world_table_titles)

    else:
        print("Không tìm thấy bảng với id 'data_content_onday'.")
        df = pd.DataFrame()  # Trả về DataFrame rỗng nếu không tìm thấy bảng
        send_slack_notification('[ERROR] Chương trình chạy thất bại', webhook_url)

    return df


# 2.2 Lưu và xứ lý dữ liệu vào database PostgreSQL
def create_and_connect_to_database(db_name, user, password, host='localhost', port='5432'):
    """Tạo một database mới nếu chưa tồn tại và kết nối đến nó."""
    try:
        # Kết nối đến PostgreSQL (mặc định là database hệ thống 'postgres')
        engine = create_engine(f'postgresql://{user}:{password}@{host}:{port}', isolation_level='AUTOCOMMIT')

        # Tạo database nếu chưa tồn tại
        with engine.connect() as conn:
            result = conn.execute(
                text("SELECT 1 FROM pg_catalog.pg_database WHERE datname = :db_name"),
                {"db_name": db_name}
            )
            exists = result.fetchone()
            if not exists:
                conn.execute(text(f"CREATE DATABASE {db_name}"))
                print(f"Database '{db_name}' đã được tạo.")
            else:
                print(f"Database '{db_name}' đã tồn tại.")

        # Kết nối đến database vừa tạo
        engine = create_engine(f'postgresql://{user}:{password}@{host}:{port}/{db_name}')
        print(f"Kết nối thành công đến database: {db_name}")
        return engine

    except Exception as e:
        print(f"Lỗi khi tạo hoặc kết nối đến database: {e}")
        return None


def upload_excel_to_postgres(db_config, company):
    try:
        # Tìm tất cả các file Excel với pattern data_thue_dien_tu*.xlsx
        list_of_files = glob.glob(f"./data_thue_dien_tu*.xlsx")
        if not list_of_files:
            print("Không tìm thấy file Excel nào trong thư mục.")
            return

        # Sắp xếp files theo thời gian tạo, lấy file mới nhất
        latest_file = max(list_of_files, key=os.path.getctime)
        print(f"Sử dụng file mới nhất: {latest_file}")

        # Kết nối tới PostgreSQL
        engine = create_engine(
            f"postgresql+psycopg2://{db_config['user']}:{db_config['password']}@{db_config['host']}:{db_config['port']}/{db_config['database']}"
        )

        # Đọc file Excel mới nhất
        data = pd.read_excel(latest_file)

        # Đổi tên cột và loại bỏ cột `stt`
        data.columns = [
            "thu_tu_thanh_toan", "co_quan_thu", "loai_nghia_vu", 
            "so_tham_chieu", "id_khoan_phai_nop", "so_quyet_dinh_so_thong_bao", 
            "ky_thue", "ngay_quyet_dinh", "tieu_muc", "so_tien", 
            "loai_tien", "ma_chuong", "dbhc", "han_nop_ngay", 
            "so_tien_da_nop", "trang_thai", "tinh_chat_khoan_nop"
        ]

        # Thay thế NaN bằng chuỗi rỗng
        data.fillna('', inplace=True)

        # Thêm cột mới
        data['created_at'] = pd.to_datetime('now')
        data['company'] = company

        # Tạo bảng data_thuedt nếu chưa tồn tại
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS data_thuedt (
                    id SERIAL PRIMARY KEY,
                    thu_tu_thanh_toan VARCHAR,
                    co_quan_thu VARCHAR,
                    loai_nghia_vu VARCHAR,
                    so_tham_chieu VARCHAR,
                    id_khoan_phai_nop VARCHAR NOT NULL,
                    so_quyet_dinh_so_thong_bao VARCHAR,
                    ky_thue VARCHAR,
                    ngay_quyet_dinh VARCHAR,
                    tieu_muc VARCHAR,
                    so_tien VARCHAR,
                    loai_tien VARCHAR,
                    ma_chuong VARCHAR,
                    dbhc VARCHAR,
                    han_nop_ngay VARCHAR,
                    so_tien_da_nop VARCHAR,
                    trang_thai VARCHAR,
                    tinh_chat_khoan_nop VARCHAR,
                    created_at TIMESTAMP,
                    company VARCHAR NOT NULL,
                    UNIQUE (id_khoan_phai_nop, company)
                );
            """))
            print("Đã kiểm tra và tạo bảng data_thuedt trong database.")

            # Kiểm tra và tạo khóa ngoại company nếu chưa tồn tại
            try:
                conn.execute(text("""
                    ALTER TABLE data_thuedt
                    ADD CONSTRAINT fk_company FOREIGN KEY (company) 
                    REFERENCES company_information (company) ON DELETE CASCADE;
                """))
                print("Đã tạo khóa ngoại company trong bảng data_thuedt.")
            except Exception as e:
                if "already exists" in str(e):
                    print("Khóa ngoại company đã tồn tại trong bảng data_thuedt.")
                else:
                    raise e

        # Lưu dữ liệu vào bảng, xử lý trùng lặp
        with engine.begin() as conn:
            for _, row in data.iterrows():
                row_data = row.to_dict()
                conn.execute(text("""
                    INSERT INTO data_thuedt (
                        thu_tu_thanh_toan, co_quan_thu, loai_nghia_vu, 
                        so_tham_chieu, id_khoan_phai_nop, so_quyet_dinh_so_thong_bao, 
                        ky_thue, ngay_quyet_dinh, tieu_muc, so_tien, 
                        loai_tien, ma_chuong, dbhc, han_nop_ngay, 
                        so_tien_da_nop, trang_thai, tinh_chat_khoan_nop, 
                        created_at, company
                    )
                    VALUES (
                        :thu_tu_thanh_toan, :co_quan_thu, :loai_nghia_vu, 
                        :so_tham_chieu, :id_khoan_phai_nop, :so_quyet_dinh_so_thong_bao, 
                        :ky_thue, :ngay_quyet_dinh, :tieu_muc, :so_tien, 
                        :loai_tien, :ma_chuong, :dbhc, :han_nop_ngay, 
                        :so_tien_da_nop, :trang_thai, :tinh_chat_khoan_nop, 
                        :created_at, :company
                    )
                    ON CONFLICT (id_khoan_phai_nop, company) 
                    DO UPDATE SET 
                        thu_tu_thanh_toan = EXCLUDED.thu_tu_thanh_toan,
                        co_quan_thu = EXCLUDED.co_quan_thu,
                        loai_nghia_vu = EXCLUDED.loai_nghia_vu,
                        so_tham_chieu = EXCLUDED.so_tham_chieu,
                        so_quyet_dinh_so_thong_bao = EXCLUDED.so_quyet_dinh_so_thong_bao,
                        ky_thue = EXCLUDED.ky_thue,
                        ngay_quyet_dinh = EXCLUDED.ngay_quyet_dinh,
                        tieu_muc = EXCLUDED.tieu_muc,
                        so_tien = EXCLUDED.so_tien,
                        loai_tien = EXCLUDED.loai_tien,
                        ma_chuong = EXCLUDED.ma_chuong,
                        dbhc = EXCLUDED.dbhc,
                        han_nop_ngay = EXCLUDED.han_nop_ngay,
                        so_tien_da_nop = EXCLUDED.so_tien_da_nop,
                        trang_thai = EXCLUDED.trang_thai,
                        tinh_chat_khoan_nop = EXCLUDED.tinh_chat_khoan_nop,
                        created_at = EXCLUDED.created_at;
                """), row_data)

        print("Đã tải dữ liệu vào bảng data_thuedt thành công.")

    except Exception as e:
        print(f"Lỗi khi tải dữ liệu lên PostgreSQL: {e}")
        raise e


# Hàm lấy dữ liệu từ bảng company_information
def fetch_company_information(engine):
    query = text("SELECT company, thue_username, thue_password FROM company_information;")
    try:
        with engine.connect() as conn:
            result = conn.execute(query)
            rows = result.fetchall()

            # Lọc các công ty không có thue_username hoặc thue_password
            filtered_rows = [
                {"company": row[0], "thue_username": row[1], "thue_password": row[2]} 
                for row in rows 
                if row[1] and row[2] 
            ]
           
            return filtered_rows 
    except Exception as e:
        print(f"Error fetching data from 'company_information': {e}")
        return []

def clean_data(directory_path=".", file_extensions=(".xlsx")):
    """
    Xóa tất cả các file dữ liệu trong thư mục được chỉ định có phần mở rộng cụ thể.
    
    Args:
        directory_path (str): Đường dẫn đến thư mục chứa file dữ liệu. Mặc định là thư mục hiện tại.
        file_extensions (tuple): Các phần mở rộng của file cần xóa. Mặc định là ('.csv', '.pdf').
    
    Returns:
        None
    """
    try:
        files_removed = 0
        for file_name in os.listdir(directory_path):
            if file_name.endswith(file_extensions):
                file_path = os.path.join(directory_path, file_name)
                os.remove(file_path)
                files_removed += 1
                print(f"[INFO] Đã xóa file: {file_path}")
        
        if files_removed == 0:
            print(f"[INFO] Không có file nào với đuôi {file_extensions} trong thư mục '{directory_path}' để xóa.")
        else:
            print(f"[INFO] Tổng số file đã xóa: {files_removed}")

    except Exception as e:
        print(f"[ERROR] Lỗi khi xóa dữ liệu: {e}")


def main():
    """Main function to run the crawler with parsed arguments."""
    # Parse command line arguments
    args = parse_arguments()
    
    # Initialize database configuration
    db_config = {
        'host': args.db_host,
        'port': args.db_port,
        'user': args.db_user,
        'password': args.db_password,
        'database': args.db_name
    }

    engine = create_and_connect_to_database(args.db_name, args.db_user, args.db_password, args.db_host, args.db_port)
    
    # Lấy danh sách công ty từ cơ sở dữ liệu
    company_data_list = fetch_company_information(engine)
    
    # Kiểm tra xem có dữ liệu công ty hay không
    if not company_data_list:
        print("Không có công ty nào để xử lý. Kết thúc chương trình.")
        return

    total_companies = len(company_data_list)
    print(f"\nTổng số công ty cần xử lý: {total_companies}")
    
    # Initialize webdriver
    driver = initialize_driver()
    captcha_image_path = "captcha_image.png"

    successful_companies = []
    failed_companies = []

    # Xử lý từng công ty
    for idx, company_data in enumerate(company_data_list, start=1):
        company, username, password = (
            company_data["company"],
            company_data["thue_username"],
            company_data["thue_password"],
        )
        
        # Kiểm tra tính hợp lệ của dữ liệu công ty
        if not (company and username and password):
            print(f"Dữ liệu công ty không hợp lệ cho công ty thứ {idx}. Kết thúc chương trình.")
            driver.quit()
            return
        
        print(f"\nĐang xử lý công ty thứ {idx}/{total_companies}: {company}")

        try:
            # Mở tab mới
            driver.execute_script("window.open('');")
            new_tab = driver.window_handles[-1]
            driver.switch_to.window(new_tab)

            # Đăng nhập
            login_to_thuedientu(driver, username, password, company)
            save_captcha_image(driver)
            enter_verification_code(driver, captcha_image_path)
            submit_form(driver, username, password, captcha_image_path)

            # Crawling dữ liệu
            df = crawl(driver)

            if engine and not df.empty:
                # Lưu vào file Excel
                file_path = f'data_thue_dien_tu.xlsx'
                unique_file_name = save_to_excel_with_style(df, file_path)
                adjust_column_width(unique_file_name)

                # Tải dữ liệu lên cơ sở dữ liệu
                upload_excel_to_postgres(db_config, company)
                successful_companies.append(company)  # Thêm vào danh sách thành công
            else:
                failed_companies.append(company)  # Thêm vào danh sách thất bại

        except Exception as e:
            print(f"Đã xảy ra lỗi: {e}")
            failed_companies.append(company)  # Thêm vào danh sách thất bại
            send_slack_notification(f"Lỗi khi xử lý công ty {company}: {e}", args.webhook_url)

    # Đóng tất cả các tab sau khi hoàn tất
    driver.quit()  # Đóng WebDriver sau khi xử lý tất cả công ty
    clean_data(directory_path=".", file_extensions=(".xlsx"))
    # In báo cáo tổng kết
    print("\n=========== Báo cáo tổng kết ===========")
    print(f"Số công ty cần chạy: {total_companies}")
    print(f"Số công ty chạy thành công: {len(successful_companies)}")
    print(f"Số công ty chạy thất bại: {len(failed_companies)}")

    send_slack_notification("\n=========== Báo cáo tổng kết ===========", webhook_url)
    send_slack_notification(f"Số công ty cần chạy: {total_companies}", webhook_url)
    send_slack_notification(f"Số công ty chạy thành công: {len(successful_companies)}", webhook_url)
    send_slack_notification(f"Số công ty chạy thất bại: {len(failed_companies)}", webhook_url)
    
    if successful_companies:
        print("- Công ty chạy thành công:")
        send_slack_notification("- Công ty chạy thành công:", webhook_url)
        for company in successful_companies:
            print(f" {company}")
            send_slack_notification(f" {company}", webhook_url)

    if failed_companies:
        print("- Công ty chạy thất bại:")
        send_slack_notification("- Công ty chạy thất bại:", webhook_url)
        for company in failed_companies:
            print(f" {company}")
            send_slack_notification(f" {company}", webhook_url)

if __name__ == '__main__':
    main()
